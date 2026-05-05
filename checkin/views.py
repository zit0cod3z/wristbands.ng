import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db import models
from events.models import Event
from registrations.models import Registration
from .models import CheckInLog
from .god_mode_utils import safe_user


def is_admin(user):
    from accounts.god_mode import is_god_mode
    return user.is_authenticated and (
        is_god_mode(user) or user.is_staff or user.is_superuser or hasattr(user, 'admin_profile')
    )


def _can_access_event(user, event):
    from accounts.god_mode import is_god_mode
    if is_god_mode(user) or user.is_superuser or user.is_staff:
        return True
    if hasattr(user, 'admin_profile'):
        return user.admin_profile.can_access_event(event)
    return False


def _get_accessible_events(user):
    from accounts.god_mode import is_god_mode
    if is_god_mode(user) or user.is_superuser or user.is_staff:
        return Event.objects.all()
    if hasattr(user, 'admin_profile'):
        return user.admin_profile.get_accessible_events()
    return Event.objects.none()


# ─────────────────────────────────────────────
# 1. Central overview – all events + their check-in stats
# ─────────────────────────────────────────────
@login_required
def event_select(request):
    from django.db.models import Count, Q
    events = _get_accessible_events(request.user).filter(status='published').order_by('-start_date').annotate(
        total_regs=Count('registrations', filter=Q(registrations__status='confirmed')),
        total_checkins=Count('registrations', filter=Q(
            registrations__status='confirmed', registrations__checked_in=True
        )),
    )
    return render(request, 'checkin/event_select.html', {'events': events})


# ─────────────────────────────────────────────
# 2. Scanner interface  (PDA / smartphone)
# ─────────────────────────────────────────────
@login_required
def scanner(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'You do not have permission to scan this event.')
        return redirect('checkin:event_select')
    total = event.registrations.filter(status='confirmed').count()
    checked_in = event.registrations.filter(status='confirmed', checked_in=True).count()
    return render(request, 'checkin/scanner.html', {
        'event': event,
        'total': total,
        'checked_in': checked_in,
    })


# ─────────────────────────────────────────────
# 3. Process scan  (AJAX POST from scanner page)
# ─────────────────────────────────────────────
@login_required
@require_POST
def process_scan(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    try:
        body = json.loads(request.body)
        raw_code = body.get('code', '').strip()
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Invalid request body — expected JSON'}, status=400)

    if not raw_code:
        return JsonResponse({'status': 'error', 'message': 'No code received'}, status=400)

    device_info = request.META.get('HTTP_USER_AGENT', '')[:255]
    registration = None
    success = False
    message = ''
    reg_data = {}

    # ── QR Code Lookup — 4 formats supported ─────────────────────────────
    # 1. WristbandsNG native:  WRISTBANDSNG|<uuid>|<reg_code>|<event_title>
    # 2. External QR ID:       exact string stored in registration.external_qr_id
    # 3. EXT- prefixed code:   EXT-<external_qr_id> (auto-generated for imports)
    # 4. Plain registration code: WRI-XXXXXXX

    if raw_code.startswith('WRISTBANDSNG|') or raw_code.startswith('EVENTPRO|'):
        # Native WristbandsNG QR
        parts = raw_code.split('|')
        if len(parts) >= 3:
            reg_id   = parts[1].strip()
            reg_code = parts[2].strip()
            try:
                registration = Registration.objects.get(pk=reg_id, event=event)
            except Registration.DoesNotExist:
                try:
                    registration = Registration.objects.get(
                        registration_code=reg_code, event=event
                    )
                except Registration.DoesNotExist:
                    registration = None
    else:
        # Strategy 1: exact external QR ID match (handles any external system QR)
        registration = Registration.objects.filter(
            external_qr_id=raw_code, event=event
        ).first()

        # Strategy 2: plain WristbandsNG registration code (WRI-XXXXXXX)
        if not registration:
            try:
                registration = Registration.objects.get(
                    registration_code=raw_code, event=event
                )
            except Registration.DoesNotExist:
                registration = None

        # Strategy 3: case-insensitive external QR ID
        if not registration:
            registration = Registration.objects.filter(
                external_qr_id__iexact=raw_code, event=event
            ).first()

        # Strategy 4: strip EXT- prefix if someone types the registration_code
        # e.g. scanning "EXT-51" should find external_qr_id="51"
        if not registration and raw_code.upper().startswith('EXT-'):
            stripped = raw_code[4:].strip()
            registration = Registration.objects.filter(
                external_qr_id=stripped, event=event
            ).first()
            if not registration:
                registration = Registration.objects.filter(
                    external_qr_id__iexact=stripped, event=event
                ).first()

    if registration is None:
        message = 'QR code not found for this event.'
        status_key = 'not_found'
    elif registration.status != 'confirmed':
        message = f'Registration is {registration.get_status_display()} — entry denied.'
        status_key = 'denied'
    elif not event.checkin_open:
        # Check-in is closed — block new entries but preserve all existing data
        message = 'Check-in for this event is currently closed by the organiser.'
        status_key = 'denied'
    elif registration.checked_in:
        message = f'Already checked in at {registration.checked_in_at.strftime("%H:%M")}'
        status_key = 'duplicate'
        reg_data = _reg_payload(registration)
    else:
        registration.checked_in = True
        registration.checked_in_at = timezone.now()
        registration.save(update_fields=['checked_in', 'checked_in_at'])
        success = True
        message = f'Welcome, {registration.name}!'
        status_key = 'success'
        reg_data = _reg_payload(registration)

    # Log the scan
    CheckInLog.objects.create(
        registration=registration,
        event=event,
        scanned_code=raw_code[:500],
        scanned_by=safe_user(request.user),
        success=success,
        message=message,
        device_info=device_info,
    )

    # Broadcast to all connected dashboards/scanners via WebSocket
    payload = {
        'type': status_key,
        'message': message,
        'registration': reg_data,
        'checked_in_total': event.registrations.filter(
            status='confirmed', checked_in=True
        ).count(),
        'total': event.registrations.filter(status='confirmed').count(),
    }
    _broadcast(event_pk, payload)

    return JsonResponse({'status': status_key, 'message': message, 'data': reg_data})


# ─────────────────────────────────────────────
# 4. Check-in dashboard  (admin view)
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def dashboard(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'You do not have permission to view this event.')
        return redirect('checkin:event_select')
    checked_in = event.registrations.filter(
        status='confirmed', checked_in=True
    ).order_by('-checked_in_at')
    not_arrived = event.registrations.filter(
        status='confirmed', checked_in=False
    ).order_by('name')
    logs = CheckInLog.objects.filter(event=event).select_related('registration')[:50]
    total = event.registrations.filter(status='confirmed').count()
    return render(request, 'checkin/dashboard.html', {
        'event': event,
        'checked_in': checked_in,
        'not_arrived': not_arrived,
        'logs': logs,
        'total': total,
        'checked_in_count': checked_in.count(),
    })


# ─────────────────────────────────────────────
# 5. Live stats endpoint  (polled by dashboard)
# ─────────────────────────────────────────────
@login_required
def live_stats(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    total = event.registrations.filter(status='confirmed').count()
    checked_in = event.registrations.filter(status='confirmed', checked_in=True).count()
    recent = (
        CheckInLog.objects
        .filter(event=event, success=True)
        .select_related('registration')
        .order_by('-scanned_at')[:10]
    )
    recent_data = [
        {
            'name': log.registration.name if log.registration else '—',
            'code': log.registration.registration_code if log.registration else '—',
            'time': log.scanned_at.strftime('%H:%M:%S'),
        }
        for log in recent
    ]
    return JsonResponse({
        'total': total,
        'checked_in': checked_in,
        'remaining': total - checked_in,
        'percent': round((checked_in / total * 100) if total else 0, 1),
        'recent': recent_data,
    })


# ─────────────────────────────────────────────
# 6. Manual check-in toggle  (admin)
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
@require_POST
def manual_checkin(request, event_pk, reg_id):
    event = get_object_or_404(Event, pk=event_pk)
    reg = get_object_or_404(Registration, pk=reg_id, event=event)
    if reg.checked_in:
        reg.checked_in = False
        reg.checked_in_at = None
        reg.save(update_fields=['checked_in', 'checked_in_at'])
        msg = f'{reg.name} unchecked.'
    else:
        reg.checked_in = True
        reg.checked_in_at = timezone.now()
        reg.save(update_fields=['checked_in', 'checked_in_at'])
        msg = f'{reg.name} checked in.'
        payload = {
            'type': 'success',
            'message': msg,
            'registration': _reg_payload(reg),
            'checked_in_total': event.registrations.filter(
                status='confirmed', checked_in=True
            ).count(),
            'total': event.registrations.filter(status='confirmed').count(),
        }
        _broadcast(event_pk, payload)
    return JsonResponse({'ok': True, 'message': msg, 'checked_in': reg.checked_in})


# ─────────────────────────────────────────────
# 7. Export checked-in guests to Excel
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def export_checkin(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'Permission denied.')
        return redirect('checkin:event_select')
    registrations = (
        event.registrations
        .filter(status='confirmed', checked_in=True)
        .prefetch_related('field_data')
        .order_by('checked_in_at')
    )
    fields = list(event.form_fields.values_list('field_key', 'label'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{event.title[:25]} Check-in'

    hfill  = PatternFill(start_color='AC2376', end_color='AC2376', fill_type='solid')
    hfont  = Font(color='FFFFFF', bold=True, size=11)
    halign = Alignment(horizontal='center', vertical='center')

    base_headers = ['#', 'Name', 'Email', 'Reg Code', 'Checked In At']
    extra_headers = [label for _, label in fields]
    all_headers = base_headers + extra_headers

    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = halign
        ws.column_dimensions[cell.column_letter].width = max(16, len(h) + 4)
    ws.row_dimensions[1].height = 26

    for ri, reg in enumerate(registrations, 2):
        fmap = {d.field_key: d.value for d in reg.field_data.all()}
        row = [
            ri - 1, reg.name, reg.email, reg.registration_code,
            reg.checked_in_at.strftime('%Y-%m-%d %H:%M') if reg.checked_in_at else '',
        ] + [fmap.get(k, '') for k, _ in fields]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical='center')
            if ri % 2 == 0:
                cell.fill = PatternFill(start_color='FDF0F7', end_color='FDF0F7', fill_type='solid')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe = event.title.replace(' ', '_')[:28]
    response['Content-Disposition'] = f'attachment; filename="{safe}_checkin.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# 8. All-events overview dashboard
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def overview(request):
    from django.db.models import Count, Q
    events = _get_accessible_events(request.user).annotate(
        total_regs=Count('registrations', filter=Q(registrations__status='confirmed')),
        total_checkins=Count('registrations', filter=Q(
            registrations__status='confirmed', registrations__checked_in=True
        )),
    ).order_by('-start_date')

    # Grand totals
    grand_total    = sum(e.total_regs for e in events)
    grand_checkins = sum(e.total_checkins for e in events)

    return render(request, 'checkin/overview.html', {
        'events': events,
        'grand_total': grand_total,
        'grand_checkins': grand_checkins,
        'grand_remaining': grand_total - grand_checkins,
        'grand_pct': round(grand_checkins / grand_total * 100, 1) if grand_total else 0,
    })


# ─────────────────────────────────────────────
# 9. Export ALL events check-in data (one sheet per event)
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def export_all_checkins(request):
    from django.db.models import Q
    events = Event.objects.filter(
        registrations__status='confirmed',
        registrations__checked_in=True
    ).distinct().order_by('start_date')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    hfill  = PatternFill(start_color='AC2376', end_color='AC2376', fill_type='solid')
    hfont  = Font(color='FFFFFF', bold=True, size=11)
    halign = Alignment(horizontal='center', vertical='center')

    for event in events:
        registrations = (
            event.registrations
            .filter(status='confirmed', checked_in=True)
            .prefetch_related('field_data')
            .order_by('checked_in_at')
        )
        fields = list(event.form_fields.values_list('field_key', 'label'))

        # Sheet name: max 31 chars, no special chars
        safe_title = ''.join(c for c in event.title if c.isalnum() or c in ' _-')[:28]
        ws = wb.create_sheet(title=safe_title or f'Event_{str(event.pk)[:6]}')

        # Event info header rows
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = event.title
        title_cell.font  = Font(color='FFFFFF', bold=True, size=13)
        title_cell.fill  = PatternFill(start_color='AC2376', end_color='AC2376', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:F2')
        info_cell = ws['A2']
        info_cell.value = (
            f"Date: {event.start_date.strftime('%d %b %Y')}  |  "
            f"Venue: {event.venue}  |  "
            f"Checked In: {registrations.count()} / {event.registrations.filter(status='confirmed').count()}"
        )
        info_cell.font  = Font(color='E6573F', bold=True, size=10)
        info_cell.fill  = PatternFill(start_color='1A1014', end_color='1A1014', fill_type='solid')
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20

        # Column headers (row 3)
        base_headers = ['#', 'Name', 'Email', 'Reg Code', 'Checked In At']
        extra_headers = [label for _, label in fields]
        all_headers = base_headers + extra_headers

        for ci, h in enumerate(all_headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.fill = hfill; cell.font = hfont; cell.alignment = halign
            ws.column_dimensions[cell.column_letter].width = max(16, len(h) + 4)
        ws.row_dimensions[3].height = 24

        # Data rows (start at row 4)
        for ri, reg in enumerate(registrations, 4):
            fmap = {d.field_key: d.value for d in reg.field_data.all()}
            row_data = [
                ri - 3, reg.name, reg.email, reg.registration_code,
                reg.checked_in_at.strftime('%Y-%m-%d %H:%M') if reg.checked_in_at else '',
            ] + [fmap.get(k, '') for k, _ in fields]

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical='center')
                if ri % 2 == 0:
                    cell.fill = PatternFill(start_color='FDF0F7', end_color='FDF0F7', fill_type='solid')

    if not wb.sheetnames:
        ws = wb.create_sheet('No Data')
        ws['A1'] = 'No check-in data found across any events.'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from django.utils import timezone as tz
    stamp = tz.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="all_events_checkin_{stamp}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def _reg_payload(reg):
    return {
        'id': str(reg.id),
        'name': reg.name,
        'email': reg.email,
        'code': reg.registration_code,
        'checked_in_at': reg.checked_in_at.strftime('%H:%M:%S') if reg.checked_in_at else None,
    }


def _broadcast(event_pk, data):
    try:
        layer = get_channel_layer()
        if layer:
            async_to_sync(layer.group_send)(
                f'checkin_{event_pk}',
                {'type': 'checkin.update', 'data': data}
            )
    except Exception:
        pass  # Gracefully degrade if no channel layer configured


# ─────────────────────────────────────────────
# 10. PWA Scanner page  (installable, offline-capable)
# ─────────────────────────────────────────────
@login_required
def pwa_scanner(request, event_pk):
    import json as _json
    event = get_object_or_404(Event, pk=event_pk)
    total = event.registrations.filter(status='confirmed').count()
    checked_in = event.registrations.filter(status='confirmed', checked_in=True).count()

    # Pre-load ALL registrations into the page so offline cache
    # is populated immediately on page load — no separate fetch needed
    regs_qs = event.registrations.filter(status='confirmed').values(
        'id', 'registration_code', 'name', 'email', 'checked_in', 'checked_in_at',
        'external_qr_id'
    )
    preloaded = []
    for r in regs_qs:
        preloaded.append({
            'id':             str(r['id']),
            'code':           r['registration_code'],
            'external_qr_id': r['external_qr_id'] or '',
            'name':           r['name'],
            'email':          r['email'],
            'checked_in':     r['checked_in'],
            'checked_in_at':  r['checked_in_at'].strftime('%H:%M') if r['checked_in_at'] else None,
        })

    return render(request, 'checkin/pwa_scanner.html', {
        'event': event,
        'total': total,
        'checked_in': checked_in,
        'preloaded_regs_json': _json.dumps(preloaded),
        'preloaded_count': len(preloaded),
    })


# ─────────────────────────────────────────────
# 11. Offline registrations payload
#     Called once when online to cache all reg codes locally on the device
# ─────────────────────────────────────────────
@login_required
def offline_registrations(request, event_pk):
    """
    Returns a lightweight JSON list of all confirmed registrations for an event.
    The PWA stores this in IndexedDB so scanning works without internet.
    Only returns: id, registration_code, name, email, checked_in status.
    No sensitive extra field data is included.
    """
    event = get_object_or_404(Event, pk=event_pk)
    regs = event.registrations.filter(status='confirmed').values(
        'id', 'registration_code', 'name', 'email', 'checked_in', 'checked_in_at',
        'external_qr_id'
    )
    data = []
    for r in regs:
        data.append({
            'id':             str(r['id']),
            'code':           r['registration_code'],
            'external_qr_id': r['external_qr_id'] or '',
            'name':           r['name'],
            'email':          r['email'],
            'checked_in':     r['checked_in'],
            'checked_in_at':  r['checked_in_at'].strftime('%H:%M') if r['checked_in_at'] else None,
        })
    return JsonResponse({
        'event_id': str(event_pk),
        'event_title': event.title,
        'cached_at': timezone.now().isoformat(),
        'registrations': data,
        'total': len(data),
    })


# ─────────────────────────────────────────────
# 12. Offline sync – bulk flush queued check-ins
#     Called by PWA when it comes back online
# ─────────────────────────────────────────────
@login_required
@require_POST
def sync_offline_checkins(request, event_pk):
    """
    Accepts a JSON array of {code, scanned_at} objects that were
    scanned offline and queued in IndexedDB on the device.
    Processes each one exactly like a normal scan.
    Returns per-item results so the PWA can clear its queue.
    """
    event = get_object_or_404(Event, pk=event_pk)
    try:
        body = json.loads(request.body)
        items = body.get('queue', [])
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    results = []
    for item in items:
        raw_code = item.get('code', '').strip()
        offline_time = item.get('scanned_at', '')
        registration = None

        if raw_code.startswith('WRISTBANDSNG|') or raw_code.startswith('EVENTPRO|'):
            parts = raw_code.split('|')
            if len(parts) >= 3:
                try:
                    registration = Registration.objects.get(pk=parts[1].strip(), event=event)
                except Registration.DoesNotExist:
                    try:
                        registration = Registration.objects.get(
                            registration_code=parts[2].strip(), event=event
                        )
                    except Registration.DoesNotExist:
                        pass
        else:
            try:
                registration = Registration.objects.get(registration_code=raw_code, event=event)
            except Registration.DoesNotExist:
                pass

        if registration is None:
            status_key = 'not_found'
            message = 'Not found'
            success = False
        elif registration.status != 'confirmed':
            status_key = 'denied'
            message = f'Registration {registration.get_status_display()}'
            success = False
        elif registration.checked_in:
            status_key = 'duplicate'
            message = f'Already checked in'
            success = False
        else:
            registration.checked_in = True
            registration.checked_in_at = timezone.now()
            registration.save(update_fields=['checked_in', 'checked_in_at'])
            status_key = 'success'
            message = f'Checked in: {registration.name}'
            success = True

        CheckInLog.objects.create(
            registration=registration,
            event=event,
            scanned_code=raw_code[:500],
            scanned_by=safe_user(request.user),
            success=success,
            message=f'[OFFLINE sync] {message}',
            device_info=request.META.get('HTTP_USER_AGENT', '')[:255],
        )

        results.append({
            'code': raw_code,
            'status': status_key,
            'message': message,
            'offline_time': offline_time,
        })

    # Broadcast updated totals after bulk sync
    ci_total = event.registrations.filter(status='confirmed', checked_in=True).count()
    total = event.registrations.filter(status='confirmed').count()
    _broadcast(event_pk, {
        'type': 'sync',
        'message': f'Offline sync: {len(results)} items processed',
        'checked_in_total': ci_total,
        'total': total,
    })

    return JsonResponse({
        'synced': len(results),
        'results': results,
        'checked_in_total': ci_total,
        'total': total,
    })


# ─────────────────────────────────────────────
# 13. PWA Web App Manifest  (makes it installable)
# ─────────────────────────────────────────────
def pwa_manifest(request):
    manifest = {
        "name": "WristbandsNG Scanner",
        "short_name": "WBNG Scanner",
        "description": "Scan QR codes to check in event guests — works offline",
        "start_url": "/checkin/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#0d0d12",
        "theme_color": "#ac2376",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/static/img/icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": "/static/img/icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "maskable"
            },
            {
                "src": "/static/img/icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": "/static/img/icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "maskable"
            }
        ],
        "categories": ["utilities", "productivity"],
        "shortcuts": [
            {
                "name": "Check-in Overview",
                "url": "/checkin/overview/",
                "description": "View all events check-in status"
            },
            {
                "name": "Offline Register",
                "url": "/checkin/",
                "description": "Register guests offline"
            }
        ],
        "prefer_related_applications": False,
    }
    return JsonResponse(manifest, content_type='application/manifest+json')


# ─────────────────────────────────────────────
# 14. Service Worker  (served from root scope)
# ─────────────────────────────────────────────
def service_worker(request):
    from django.http import FileResponse
    import os
    from django.conf import settings
    sw_path = os.path.join(settings.BASE_DIR, 'static', 'js', 'sw.js')
    response = FileResponse(open(sw_path, 'rb'), content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    return response


# ─────────────────────────────────────────────
# 15. Scanner URL QR code image
#     Generates a QR code of the scanner page URL itself
#     so staff can scan it with any phone to open the scanner
# ─────────────────────────────────────────────
@login_required
def scanner_url_qr(request, event_pk):
    """Returns a PNG QR code image whose data is the full scanner URL."""
    import io
    import qrcode as _qrcode
    event = get_object_or_404(Event, pk=event_pk)
    scanner_url = request.build_absolute_uri(
        f'/checkin/{event_pk}/pwa/'
    )
    qr = _qrcode.QRCode(
        version=2,
        error_correction=_qrcode.constants.ERROR_CORRECT_H,
        box_size=8,
        border=3,
    )
    qr.add_data(scanner_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#ac2376', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return HttpResponse(buf.read(), content_type='image/png')


# ─────────────────────────────────────────────
# 16. Import guest list from Excel
#     Reads Name + Email columns, creates confirmed registrations,
#     generates QR codes, optionally marks them checked-in immediately
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def import_guest_list(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        mark_checked_in = request.POST.get('mark_checked_in') == '1'

        if not excel_file:
            from django.contrib import messages
            messages.error(request, 'Please select an Excel file.')
            return render(request, 'checkin/import_guests.html', {'event': event})

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            # Auto-detect header row — find Name and Email columns
            headers = {}
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        val = cell.strip().lower()
                        if val in ('name', 'full name', 'fullname', 'guest name', 'attendee'):
                            headers['name'] = col_idx - 1
                            header_row = row_idx
                        elif val in ('email', 'email address', 'e-mail', 'mail'):
                            headers['email'] = col_idx - 1
                            header_row = row_idx
                        elif val in ('phone', 'phone number', 'mobile', 'tel'):
                            headers['phone'] = col_idx - 1
                if header_row:
                    break

            if 'name' not in headers:
                from django.contrib import messages
                messages.error(request, 'Could not find a "Name" column. Make sure your spreadsheet has a column labelled Name.')
                return render(request, 'checkin/import_guests.html', {'event': event})

            created = 0
            skipped = 0
            already_exists = 0
            errors = []

            from registrations.models import Registration, RegistrationData
            from registrations.views import _generate_qr, _send_confirmation_email

            rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
            for row_num, row in enumerate(rows, header_row + 1):
                if not row or all(c is None for c in row):
                    continue

                name = str(row[headers['name']]).strip() if row[headers['name']] else ''
                email = str(row[headers.get('email', -1)]).strip() if headers.get('email') is not None and headers.get('email') < len(row) and row[headers.get('email')] else ''

                if not name:
                    skipped += 1
                    continue

                # Use placeholder email if none provided
                if not email:
                    email = f'guest_{row_num}_{event_pk}@import.local'

                # Skip duplicates
                if Registration.objects.filter(event=event, email=email).exists():
                    already_exists += 1
                    continue

                try:
                    reg = Registration.objects.create(
                        event=event,
                        name=name,
                        email=email,
                        status='confirmed',
                        checked_in=mark_checked_in,
                        checked_in_at=timezone.now() if mark_checked_in else None,
                    )

                    # Save phone if column found
                    if headers.get('phone') is not None and headers['phone'] < len(row) and row[headers['phone']]:
                        RegistrationData.objects.create(
                            registration=reg,
                            field_key='phone',
                            field_label='Phone',
                            value=str(row[headers['phone']]).strip(),
                        )

                    _generate_qr(reg)

                    # Only send email if real address — fire in background thread
                    if not email.endswith('@import.local'):
                        import threading
                        threading.Thread(target=_send_confirmation_email, args=(reg,), daemon=True).start()

                    created += 1

                    # Log as checked-in if flagged
                    if mark_checked_in:
                        CheckInLog.objects.create(
                            registration=reg,
                            event=event,
                            scanned_code=f'IMPORT|{reg.registration_code}',
                            scanned_by=safe_user(request.user),
                            success=True,
                            message='Imported and pre-checked-in from Excel',
                            device_info='Excel Import',
                        )

                except Exception as e:
                    errors.append(f'Row {row_num} ({name}): {str(e)}')

            from django.contrib import messages
            messages.success(request, f'Import complete — {created} added, {already_exists} already existed, {skipped} skipped.')
            if errors:
                messages.warning(request, f'{len(errors)} row(s) had errors: ' + '; '.join(errors[:5]))

            return redirect('checkin:manual_checkin_list', event_pk=event_pk)

        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Failed to read Excel file: {str(e)}')

    return render(request, 'checkin/import_guests.html', {'event': event})


# ─────────────────────────────────────────────
# 17. Manual check-in by name search
#     Search guest list by name, check them in with one tap
# ─────────────────────────────────────────────
@login_required
def manual_checkin_list(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    query = request.GET.get('q', '').strip()
    registrations = []

    if query:
        registrations = event.registrations.filter(
            status='confirmed'
        ).filter(
            models.Q(name__icontains=query) |
            models.Q(email__icontains=query) |
            models.Q(registration_code__icontains=query)
        ).order_by('name')

    total = event.registrations.filter(status='confirmed').count()
    checked_in_count = event.registrations.filter(status='confirmed', checked_in=True).count()

    return render(request, 'checkin/manual_checkin_list.html', {
        'event': event,
        'registrations': registrations,
        'query': query,
        'total': total,
        'checked_in_count': checked_in_count,
    })


# ─────────────────────────────────────────────
# Offline registration sync
# Accepts registrations queued offline and creates them server-side
# ─────────────────────────────────────────────
@login_required
@require_POST
def sync_offline_registrations(request, event_pk):
    """
    Accepts a JSON array of registrations collected offline.
    Creates each one server-side, generates QR, sends email.
    Returns per-item results so the PWA can clear its queue.
    """
    import json as _json
    from registrations.models import Registration, RegistrationData
    from registrations.views import _generate_qr, _send_confirmation_email

    event = get_object_or_404(Event, pk=event_pk)

    try:
        body  = _json.loads(request.body)
        items = body.get('registrations', [])
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    results = []
    created = 0
    skipped = 0

    for item in items:
        name  = (item.get('name') or '').strip()
        email = (item.get('email') or '').strip()
        extra = item.get('fields', {})   # dict of field_key: value

        if not name or not email:
            results.append({'status': 'error', 'message': 'Name and email required', 'email': email})
            skipped += 1
            continue

        # Skip duplicates
        if Registration.objects.filter(event=event, email=email).exists():
            results.append({'status': 'duplicate', 'message': f'{email} already registered', 'email': email})
            skipped += 1
            continue

        try:
            reg = Registration.objects.create(
                event=event,
                name=name,
                email=email,
                status='confirmed',
                ip_address=request.META.get('REMOTE_ADDR'),
            )

            # Save any extra field data
            for field_key, value in extra.items():
                field_obj = event.form_fields.filter(field_key=field_key).first()
                if field_obj:
                    RegistrationData.objects.create(
                        registration=reg,
                        field_key=field_key,
                        field_label=field_obj.label,
                        value=str(value),
                    )

            _generate_qr(reg)
            import threading
            threading.Thread(target=_send_confirmation_email, args=(reg,), daemon=True).start()

            results.append({
                'status': 'created',
                'name': reg.name,
                'email': reg.email,
                'code': reg.registration_code,
                'offline_id': item.get('offline_id'),
            })
            created += 1

        except Exception as e:
            results.append({'status': 'error', 'message': str(e), 'email': email})
            skipped += 1

    return JsonResponse({
        'created': created,
        'skipped': skipped,
        'results': results,
        'total_registrations': event.registrations.filter(status='confirmed').count(),
    })


# ─────────────────────────────────────────────
# Offline Registration Page
# A standalone PWA page for registering guests without internet
# ─────────────────────────────────────────────
@login_required
def offline_reg_page(request, event_pk):
    import json as _json
    event = get_object_or_404(Event, pk=event_pk)
    fields = list(event.form_fields.values('field_key', 'label', 'field_type',
                                           'placeholder', 'is_required', 'options', 'order'))
    # Pre-load existing registrations so duplicates can be caught offline
    existing = list(event.registrations.filter(status='confirmed').values_list('email', flat=True))
    return render(request, 'checkin/offline_reg_page.html', {
        'event': event,
        'fields_json': _json.dumps(fields),
        'existing_emails_json': _json.dumps(existing),
        'reg_count': event.registrations.filter(status='confirmed').count(),
        'sync_url': request.build_absolute_uri(f'/checkin/{event_pk}/sync-offline-registrations/'),
    })


# ─────────────────────────────────────────────
# External QR Import
# Scan/paste any external QR code, extract its data,
# create a confirmed WristbandsNG registration with QR code
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def external_qr_import_page(request, event_pk):
    """Render the external QR import page."""
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'Permission denied.')
        return redirect('checkin:event_select')
    fields = event.form_fields.all()
    return render(request, 'checkin/external_qr_import.html', {
        'event': event,
        'fields': fields,
    })


@login_required
@login_required
@user_passes_test(is_admin)
@require_POST
def external_qr_import_process(request, event_pk):
    """
    Process an external QR code payload.
    Accepts JSON: { qr_data: "raw QR string", field_map: {field_key: value, ...} }
    Creates a confirmed registration, generates WristbandsNG QR, sends email.
    """
    import json as _json
    from registrations.models import Registration, RegistrationData
    from registrations.views import _generate_qr, _send_confirmation_email

    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        body = _json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    qr_data   = (body.get('qr_data') or '').strip()
    field_map = body.get('field_map', {})   # {field_key: value}
    name      = (field_map.get('name') or body.get('name') or '').strip()
    email     = (field_map.get('email') or body.get('email') or '').strip()
    send_email_flag = body.get('send_email', True)

    if not name:
        return JsonResponse({'error': 'Name is required to create a registration.'}, status=400)

    # Generate placeholder email if none provided
    if not email:
        import uuid as _uuid
        email = f'external_{_uuid.uuid4().hex[:8]}@import.wbng'

    # Duplicate check
    if Registration.objects.filter(event=event, email=email).exists():
        existing = Registration.objects.get(event=event, email=email)
        return JsonResponse({
            'status': 'duplicate',
            'message': f'{name} is already registered for this event.',
            'registration_code': existing.registration_code,
            'existing': True,
        })

    # Create registration — store external QR ID directly on the model
    # so the scanner can find it by scanning the original external QR code
    reg = Registration.objects.create(
        event=event,
        name=name,
        email=email,
        status='confirmed',
        ip_address=request.META.get('REMOTE_ADDR'),
        external_qr_id=qr_data[:500] if qr_data else None,
    )

    # Save all mapped fields (excluding name/email already on model)
    for field_key, value in field_map.items():
        if field_key in ('name', 'email', 'external_qr_id'):
            continue
        field_obj = event.form_fields.filter(field_key=field_key).first()
        label = field_obj.label if field_obj else field_key.replace('_', ' ').title()
        RegistrationData.objects.create(
            registration=reg,
            field_key=field_key,
            field_label=label,
            value=str(value) if value else '',
        )

    # ── QR Code strategy ──────────────────────────────────────────────────
    # External QR ID provided → use it as-is, no new QR generated.
    # The scanner matches external_qr_id when the original QR is scanned.
    # No external QR ID → generate a WristbandsNG QR code as normal.
    if not qr_data:
        _generate_qr(reg)

    # Send confirmation email if requested and email is real
    if send_email_flag and not email.endswith('@import.wbng'):
        import threading
        threading.Thread(target=_send_confirmation_email, args=(reg,), daemon=True).start()

    return JsonResponse({
        'status': 'created',
        'message': f'{name} registered successfully.',
        'registration_code': reg.registration_code,
        'name': reg.name,
        'email': reg.email,
        'qr_url': reg.qr_code.url if reg.qr_code else None,
        'email_sent': reg.email_sent,
        'qr_mode': 'external' if qr_data else 'generated',
        'external_qr_id': qr_data if qr_data else None,
    })


# ─────────────────────────────────────────────
# Bulk External QR Import from Excel/CSV
# Upload a file containing QR code IDs + guest data
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def external_qr_bulk_import(request, event_pk):
    """
    POST: Upload Excel/CSV with columns:
      qr_id (required), name (required), email, phone, + any other fields
    Creates a confirmed registration for each row, stores the external QR ID,
    generates a WristbandsNG QR code, and optionally sends confirmation emails.
    Returns JSON with per-row results.
    """
    import csv
    import io as _io
    from registrations.models import Registration, RegistrationData
    from registrations.views import _generate_qr, _send_confirmation_email
    from django.contrib import messages as _messages

    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    uploaded = request.FILES.get('import_file')
    send_emails = request.POST.get('send_emails', '1') == '1'

    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    filename = uploaded.name.lower()
    results  = []
    created  = 0
    skipped  = 0
    errors   = 0

    try:
        # ── Read file into rows ───────────────────────────────────────────
        if filename.endswith('.csv'):
            text    = uploaded.read().decode('utf-8-sig')
            reader  = csv.DictReader(_io.StringIO(text))
            rows    = list(reader)
            headers = reader.fieldnames or []
        elif filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return JsonResponse({'error': 'File is empty'}, status=400)
            headers = [str(h).strip() if h else '' for h in all_rows[0]]
            rows = []
            for row in all_rows[1:]:
                if any(cell is not None for cell in row):
                    rows.append(dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])))
        else:
            return JsonResponse({'error': 'Unsupported file type. Use .xlsx, .xls, or .csv'}, status=400)

        # ── Detect column names (flexible matching) ───────────────────────
        def find_col(candidates, cols):
            """Find first matching column name (case-insensitive, whole-word partial match)."""
            import re as _re
            cols_lower = {c.lower().strip(): c for c in cols}
            # Pass 1: exact match
            for cand in candidates:
                if cand.lower() in cols_lower:
                    return cols_lower[cand.lower()]
            # Pass 2: column name contains the full candidate as a word/phrase
            for cand in candidates:
                cand_l = cand.lower()
                for col_key, col_orig in cols_lower.items():
                    # candidate must appear as a whole token in the column name
                    # e.g. "qr" matches "qr code" but NOT "id" matching "qr_id"
                    pattern = r'(?<![a-z0-9])' + _re.escape(cand_l) + r'(?![a-z0-9])'
                    if _re.search(pattern, col_key):
                        return col_orig
            return None

        # ── QR ID column — STRICT: must explicitly contain 'qr', 'ticket',
        #    'barcode', 'code', 'ref', 'pass', 'access', 'confirmation', or 'booking'
        #    Generic terms like 'id', 'number', 'no', 'serial' are EXCLUDED
        #    to prevent picking up role IDs, row numbers, or other numeric columns.
        col_qr = find_col([
            # QR-specific (highest priority)
            'qr_id', 'qr id', 'qr code', 'qrcode', 'qr no', 'qr number', 'qr',
            'external qr', 'external qr id', 'external qr code',
            # Ticket-specific
            'ticket_id', 'ticket id', 'ticket code', 'ticket no', 'ticket number',
            'ticket ref', 'ticket reference', 'ticket',
            # Barcode
            'barcode', 'bar code', 'bar_code',
            # Code/Reference (must have these exact words — not just 'id')
            'access code', 'access_code', 'passcode', 'pass code', 'pass_code',
            'confirmation code', 'confirmation_code', 'confirmation no', 'confirmation',
            'booking code', 'booking ref', 'booking reference', 'booking_ref',
            'registration code', 'reg code', 'reg_code',
            'reference code', 'ref code', 'ref_code',
            'entry code', 'entry_code',
            'event code', 'event_code',
            'guest code', 'guest_code',
            'invite code', 'invite_code',
            'promo code', 'promo_code',
        ], headers)

        col_name  = find_col(['name', 'full name', 'fullname', 'guest name', 'attendee name', 'attendee', 'guest'], headers)
        col_email = find_col(['email', 'email address', 'e-mail', 'mail'], headers)
        col_phone = find_col(['phone', 'phone number', 'mobile', 'tel', 'telephone', 'mobile number', 'contact'], headers)

        # ── Fallback: only if NO QR column found AND admin explicitly named a column
        #    with 'code' or 'ref' anywhere in the name (still excludes plain 'id')
        if not col_qr and headers:
            skip = {col_name, col_email, col_phone}
            for h in headers:
                if not h or h in skip:
                    continue
                h_lower = h.lower().strip()
                # Only use as fallback if column name contains a meaningful keyword
                # Explicitly exclude plain 'id', 'number', 'no', 'serial', 'role'
                if any(kw in h_lower for kw in ['qr', 'ticket', 'barcode', 'code', 'ref', 'pass', 'access', 'confirm', 'booking', 'entry', 'invite']):
                    # Make sure it's not just 'id' or 'role id' or 'number'
                    if h_lower not in ('id', 'number', 'no', 'serial', 'role', 'role id', 'role_id'):
                        col_qr = h
                        break

        if not col_name:
            return JsonResponse({
                'error': 'Could not find a Name column. '
                         'Make sure your file has a column labelled "Name" or "Full Name".',
                'detected_columns': headers,
            }, status=400)

        # ── Process each row ──────────────────────────────────────────────
        for row_num, row in enumerate(rows, 2):
            name     = row.get(col_name, '').strip()  if col_name  else ''
            email    = row.get(col_email, '').strip()  if col_email else ''
            phone    = row.get(col_phone, '').strip()  if col_phone else ''
            ext_qr   = row.get(col_qr, '').strip()    if col_qr    else ''

            if not name:
                results.append({'row': row_num, 'status': 'skipped', 'reason': 'Empty name', 'name': '', 'email': email})
                skipped += 1
                continue

            # Generate placeholder email if missing
            if not email:
                import uuid as _uuid
                email = f'ext_{_uuid.uuid4().hex[:8]}@import.wbng'

            # Duplicate check by email
            if Registration.objects.filter(event=event, email=email).exists():
                existing = Registration.objects.get(event=event, email=email)
                results.append({
                    'row': row_num, 'status': 'duplicate',
                    'name': name, 'email': email,
                    'code': existing.registration_code,
                    'reason': 'Email already registered',
                })
                skipped += 1
                continue

            # Also check by external QR ID if provided
            if ext_qr:
                existing_by_qr = Registration.objects.filter(
                    external_qr_id=ext_qr,
                    event=event,
                ).first()
                if existing_by_qr:
                    results.append({
                        'row': row_num, 'status': 'duplicate',
                        'name': name, 'email': email,
                        'code': existing_by_qr.registration_code,
                        'reason': f'External QR ID "{ext_qr}" already imported',
                    })
                    skipped += 1
                    continue

            try:
                reg = Registration.objects.create(
                    event=event,
                    name=name,
                    email=email,
                    status='confirmed',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    external_qr_id=ext_qr if ext_qr else None,
                )

                # Store phone if present
                if phone:
                    RegistrationData.objects.create(
                        registration=reg,
                        field_key='phone',
                        field_label='Phone',
                        value=phone,
                    )

                # Store ALL other columns as extra data
                skip_cols = {col_name, col_email, col_phone, col_qr}
                for col, val in row.items():
                    if col in skip_cols or not col or not val:
                        continue
                    field_key = col.lower().strip().replace(' ', '_').replace('-', '_')[:50]
                    RegistrationData.objects.create(
                        registration=reg,
                        field_key=field_key,
                        field_label=col,
                        value=str(val)[:500],
                    )

                # ── QR Code strategy ──────────────────────────────────────
                # If an external QR ID was provided, we DO NOT generate a new
                # WristbandsNG QR code. The external QR ID IS the check-in code.
                # Scanning the original external QR will find this registration
                # via the external_qr_id field and check the guest in directly.
                #
                # If no external QR ID, generate a WristbandsNG QR code as normal.
                if not ext_qr:
                    _generate_qr(reg)

                # Send confirmation email (only if real email address)
                if send_emails and not email.endswith('@import.wbng'):
                    import threading
                    threading.Thread(target=_send_confirmation_email, args=(reg,), daemon=True).start()

                results.append({
                    'row': row_num, 'status': 'created',
                    'name': name, 'email': email,
                    'code': reg.registration_code,
                    'external_qr_id': ext_qr,
                    'email_sent': reg.email_sent,
                    'qr_mode': 'external' if ext_qr else 'generated',
                })
                created += 1

            except Exception as e:
                results.append({
                    'row': row_num, 'status': 'error',
                    'name': name, 'email': email,
                    'reason': str(e),
                })
                errors += 1

    except Exception as e:
        return JsonResponse({'error': f'Failed to read file: {str(e)}'}, status=400)

    return JsonResponse({
        'success': True,
        'total_rows': len(rows),
        'created': created,
        'skipped': skipped,
        'errors': errors,
        'results': results,
        'detected_columns': headers,
    })


# ─────────────────────────────────────────────
# Preview Excel columns before bulk import
# Returns detected column mapping so admin can verify before importing
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
@require_POST
def external_qr_preview(request, event_pk):
    """
    Upload a file and return the detected column mapping + first 5 rows.
    Admin can verify the correct columns are detected before importing.
    """
    import csv as _csv
    import io as _io

    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    uploaded = request.FILES.get('import_file')
    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    filename = uploaded.name.lower()

    try:
        if filename.endswith('.csv'):
            text    = uploaded.read().decode('utf-8-sig')
            reader  = _csv.DictReader(_io.StringIO(text))
            rows    = list(reader)[:5]
            headers = reader.fieldnames or []
        elif filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return JsonResponse({'error': 'File is empty'}, status=400)
            headers = [str(h).strip() if h else '' for h in all_rows[0]]
            rows = []
            for row in all_rows[1:6]:
                rows.append({headers[i]: str(c).strip() if c is not None else '' for i, c in enumerate(row)})
        else:
            return JsonResponse({'error': 'Unsupported file type'}, status=400)

        # Run the same column detection logic (strict — no generic 'id', 'number', 'no')
        def find_col(candidates, cols):
            import re as _re
            cols_lower = {c.lower().strip(): c for c in cols}
            for cand in candidates:
                if cand.lower() in cols_lower:
                    return cols_lower[cand.lower()]
            for cand in candidates:
                cand_l = cand.lower()
                for col_key, col_orig in cols_lower.items():
                    pattern = r'(?<![a-z0-9])' + _re.escape(cand_l) + r'(?![a-z0-9])'
                    if _re.search(pattern, col_key):
                        return col_orig
            return None

        col_qr = find_col([
            'qr_id', 'qr id', 'qr code', 'qrcode', 'qr no', 'qr number', 'qr',
            'external qr', 'external qr id', 'external qr code',
            'ticket_id', 'ticket id', 'ticket code', 'ticket no', 'ticket number',
            'ticket ref', 'ticket reference', 'ticket',
            'barcode', 'bar code', 'bar_code',
            'access code', 'access_code', 'passcode', 'pass code',
            'confirmation code', 'confirmation_code', 'confirmation no', 'confirmation',
            'booking code', 'booking ref', 'booking reference',
            'registration code', 'reg code', 'reg_code',
            'reference code', 'ref code', 'ref_code',
            'entry code', 'entry_code',
            'event code', 'event_code',
            'guest code', 'guest_code',
            'invite code', 'invite_code',
        ], headers)

        col_name  = find_col(['name', 'full name', 'fullname', 'guest name', 'attendee name', 'attendee', 'guest'], headers)
        col_email = find_col(['email', 'email address', 'e-mail', 'mail'], headers)
        col_phone = find_col(['phone', 'phone number', 'mobile', 'tel', 'telephone'], headers)

        # Fallback: only columns that contain meaningful QR-related keywords
        if not col_qr and headers:
            skip = {col_name, col_email, col_phone}
            for h in headers:
                if not h or h in skip:
                    continue
                h_lower = h.lower().strip()
                if any(kw in h_lower for kw in ['qr', 'ticket', 'barcode', 'code', 'ref', 'pass', 'access', 'confirm', 'booking', 'entry', 'invite']):
                    if h_lower not in ('id', 'number', 'no', 'serial', 'role', 'role id', 'role_id'):
                        col_qr = h
                        break

        return JsonResponse({
            'headers': headers,
            'detected': {
                'qr_id':  col_qr,
                'name':   col_name,
                'email':  col_email,
                'phone':  col_phone,
            },
            'sample_rows': rows,
            'total_rows': len(rows),
        })

    except Exception as e:
        return JsonResponse({'error': f'Failed to read file: {str(e)}'}, status=400)
