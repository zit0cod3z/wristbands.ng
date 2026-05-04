import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from events.models import Event
from .models import Registration, RegistrationData
from registrations.views import _generate_qr, _send_confirmation_email


def is_admin(user):
    from accounts.god_mode import is_god_mode
    return user.is_authenticated and (
        is_god_mode(user) or user.is_staff or user.is_superuser
        or hasattr(user, 'admin_profile')
    )


def _get_accessible_events(user):
    from accounts.god_mode import is_god_mode
    if is_god_mode(user) or user.is_superuser or user.is_staff:
        return Event.objects.all()
    if hasattr(user, 'admin_profile'):
        return user.admin_profile.get_accessible_events()
    return Event.objects.none()


def _can_access_event(user, event):
    from accounts.god_mode import is_god_mode
    if is_god_mode(user) or user.is_superuser or user.is_staff:
        return True
    p = getattr(user, 'admin_profile', None)
    return p is not None and p.can_access_event(event)


def _can_manage_registrations(user, event):
    """Moderators are view-only — cannot add/delete/send QR."""
    from accounts.god_mode import is_god_mode
    if is_god_mode(user) or user.is_superuser or user.is_staff:
        return True
    p = getattr(user, 'admin_profile', None)
    return p is not None and p.can_manage_registrations(event)


def _can_access_event(user, event):
    if user.is_superuser or user.is_staff:
        return True
    if hasattr(user, 'admin_profile'):
        return user.admin_profile.can_access_event(event)
    return False


@login_required
@user_passes_test(is_admin)
def registrations_list(request):
    events = _get_accessible_events(request.user).order_by('-created_at')
    event_data = [
        {'event': e, 'count': e.registrations.filter(status='confirmed').count()}
        for e in events
    ]
    return render(request, 'dashboard/registrations/list.html', {'event_data': event_data})


@login_required
@user_passes_test(is_admin)
def event_registrations(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'You do not have permission to view this event.')
        return redirect('dashboard:registrations_list')
    registrations = event.registrations.prefetch_related('field_data').order_by('-registered_at')
    search = request.GET.get('q', '')
    if search:
        registrations = registrations.filter(name__icontains=search) | registrations.filter(email__icontains=search)
    return render(request, 'dashboard/registrations/event_registrations.html', {
        'event': event,
        'registrations': registrations,
        'search': search,
    })


@login_required
@user_passes_test(is_admin)
def send_qr_manual(request, reg_id):
    reg = get_object_or_404(Registration, pk=reg_id)
    if not _can_manage_registrations(request.user, reg.event):
        messages.error(request, 'Moderators cannot perform this action.')
        return redirect('dashboard:event_registrations', event_pk=reg.event.pk)
    if not reg.qr_code:
        _generate_qr(reg)
    _send_confirmation_email(reg)
    messages.success(request, f'QR code sent to {reg.email}')
    return redirect('dashboard:event_registrations', event_pk=reg.event.pk)


@login_required
@user_passes_test(is_admin)
def delete_registration(request, reg_id):
    reg = get_object_or_404(Registration, pk=reg_id)
    if not _can_manage_registrations(request.user, reg.event):
        messages.error(request, 'Moderators cannot delete registrations.')
        return redirect('dashboard:event_registrations', event_pk=reg.event.pk)
    event_pk = reg.event.pk
    if request.method == 'POST':
        reg.delete()
        messages.success(request, 'Registration deleted.')
    return redirect('dashboard:event_registrations', event_pk=event_pk)


@login_required
@user_passes_test(is_admin)
def manual_add_registration(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_manage_registrations(request.user, event):
        messages.error(request, 'Moderators cannot add registrations.')
        return redirect('dashboard:event_registrations', event_pk=event_pk)
    fields = event.form_fields.all()
    if request.method == 'POST':
        data = request.POST
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        if not name or not email:
            messages.error(request, 'Name and email are required.')
            return render(request, 'dashboard/registrations/manual_add.html', {'event': event, 'fields': fields})
        reg, created = Registration.objects.get_or_create(
            event=event, email=email,
            defaults={'name': name}
        )
        if not created:
            messages.warning(request, 'This email is already registered.')
            return redirect('dashboard:event_registrations', event_pk=event_pk)
        for field in fields:
            value = data.get(field.field_key, '')
            RegistrationData.objects.create(
                registration=reg, field_key=field.field_key,
                field_label=field.label, value=value,
            )
        _generate_qr(reg)
        _send_confirmation_email(reg)
        messages.success(request, f'{name} registered and QR sent.')
        return redirect('dashboard:event_registrations', event_pk=event_pk)
    return render(request, 'dashboard/registrations/manual_add.html', {'event': event, 'fields': fields})


@login_required
@user_passes_test(is_admin)
def export_registrations(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    if not _can_access_event(request.user, event):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:registrations_list')
    registrations = event.registrations.prefetch_related('field_data').filter(status='confirmed').order_by('registered_at')
    fields = list(event.form_fields.values_list('field_key', 'label'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = event.title[:31]

    header_fill = PatternFill(start_color='AC2376', end_color='AC2376', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    base_headers = ['#', 'Registration Code', 'Name', 'Email', 'Status', 'Registered At', 'QR Sent', 'Checked In']
    extra_headers = [label for _, label in fields]
    all_headers = base_headers + extra_headers

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 4)
    ws.row_dimensions[1].height = 25

    for row_idx, reg in enumerate(registrations, 2):
        field_map = {d.field_key: d.value for d in reg.field_data.all()}
        row_data = [
            row_idx - 1,
            reg.registration_code,
            reg.name,
            reg.email,
            reg.get_status_display(),
            reg.registered_at.strftime('%Y-%m-%d %H:%M'),
            'Yes' if reg.email_sent else 'No',
            'Yes' if reg.checked_in else 'No',
        ] + [field_map.get(key, '') for key, _ in fields]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='FDF0F7', end_color='FDF0F7', fill_type='solid')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_title = event.title.replace(' ', '_')[:30]
    response['Content-Disposition'] = f'attachment; filename="{safe_title}_registrations.xlsx"'
    wb.save(response)
    return response
